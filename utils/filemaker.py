import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import docx
import matplotlib.pyplot as plt
import numpy as np
import datetime
import time
from openpyxl.worksheet.table import Table, TableStyleInfo

APP_ROOT = os.path.dirname(os.path.abspath(__file__))   # refers to application_top
save_file1 = os.path.join(APP_ROOT, '../download/header1/data.xlsx')
save_file2 = os.path.join(APP_ROOT, '../download/header2/data.xlsx')
save_file3 = os.path.join(APP_ROOT, '../download/header3/data.xlsx')
save_file4 = os.path.join(APP_ROOT, '../download/header4/data.xlsx')

def export_data1(tracks):
    """Save and Export one file in excel worksheet and """

    wb = openpyxl.load_workbook(filename=save_file1)
    ws = wb['Sheet1']

    for row in ws.iter_rows(min_row=2, max_col=10):
        for cell in row:
            cell.value = None

    for row_index in range(len(tracks)):
        row_num = str(row_index + 2)
        row = tracks[row_index]

        ws['A' + row_num].value = row['record_id']
        ws['B' + row_num].value = row['track_id']
        ws['C' + row_num].value = row['track_label']
        ws['D' + row_num].value = row['time_start']
        ws['E' + row_num].value = row['time_end']
        ws['F' + row_num].value = row['from_address']
        ws['G' + row_num].value = row['to_address']
        ws['H' + row_num].value = row['distance']
        ws['I' + row_num].value = row['duration']
        ws['J' + row_num].value = row['max_speed']
        ws['K' + row_num].value = row['Driver']
        ws['L' + row_num].value = row['TT_number_tags']
        ws['M' + row_num].value = row['Registration_plate']
        ws['N' + row_num].value = row['Product']
        ws['O' + row_num].value = row['Parked']
        ws['P' + row_num].value = row['Idle_time']
        ws['Q' + row_num].value = row['status']
        ws['R' + row_num].value = row['zone']

    wb.save(filename=save_file1)

    return save_file1

def export_data2(tracks, track_label, start_date, end_date):
    """Save and Export one file in excel worksheet and """

    wb = openpyxl.load_workbook(filename=save_file3)
    ws = wb['Sheet1']

    ws['A1'].value = 'Utilization Report'

    for row in ws.iter_rows(min_row=2, max_col=6):
        for cell in row:
            cell.value = None

    ws['A2'].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ' Africa/Luburrbashi'
    ws['A3'].value = 'Assets: any Vehicle'
    ws['A4'].value = 'From: ' + start_date.strftime("%Y-%m-%d")
    ws['A5'].value = 'To: ' + end_date.strftime("%Y-%m-%d")
    ws['A6'].value = 'Utilization Mode: 24 hours'
    ws['A7'].value = 'Schedule: UTILIZATION REPORTS'
    ws['A8'].value = track_label

    ws['A9'].value = 'Date'
    ws['B9'].value = 'Distance'
    ws['C9'].value = 'Parked'
    ws['D9'].value = 'Parked Percent'
    ws['E9'].value = 'Driving'
    ws['F9'].value = 'Driving Percent'
    ws['G9'].value = 'Idling'
    ws['H9'].value = 'Idling Percent'
    ws['I9'].value = 'Total Hours'

    data = []

    for row_index in range(len(tracks)):
        row = tracks[row_index]

        raw_data = []        

        raw_data.append(row['time_start'][0:10])
        raw_data.append(row['distance'])

        if row["Parked"]:
            raw_data.append(row['duration'])
        else: 
            raw_data.append('00:00:00')

        if row["Parked"]:
            raw_data.append('00:00:00')
        else: 
            raw_data.append(row['duration'])

        if row["Idle_time"]:
            raw_data.append('00:00:00')
        else: 
            raw_data.append(row['duration'])

        raw_data.append(row['duration'])

        data.append(raw_data)

    delta = datetime.timedelta(days=1)

    results = {}

    data1 = []

    while start_date <= end_date:

        date_row = []
        parked_duration = "00:00:00"
        driving_duration = "00:00:00"
        idling_duration = "00:00:00"
        duration = 0
        for row in data:
            
            if start_date.strftime('%Y-%m-%d') == row[0]:
                duration += float(row[1])
                parked_duration = sum_times(parked_duration, row[2])
                driving_duration = sum_times(driving_duration, row[3])
                idling_duration = sum_times(idling_duration, row[4])
        
        driving_seconds = time.strptime(driving_duration, "%H:%M:%S")
        idling_seconds = time.strptime(idling_duration, "%H:%M:%S")

        driving_percent = datetime.timedelta(hours=driving_seconds.tm_hour,minutes=driving_seconds.tm_min,seconds=driving_seconds.tm_sec).total_seconds() / datetime.timedelta(hours=23,minutes=59,seconds=59).total_seconds() * 100
        idling_percent = datetime.timedelta(hours=idling_seconds.tm_hour,minutes=idling_seconds.tm_min,seconds=idling_seconds.tm_sec).total_seconds() / datetime.timedelta(hours=23,minutes=59,seconds=59).total_seconds() * 100
        date_row.append(round(driving_percent, 1))
        date_row.append(round(idling_percent, 1))
        results[start_date] = date_row

        data1_row = []
        data1_row.append(start_date.strftime('%Y-%m-%d'))
        data1_row.append(str(round(duration, 1)))
        if duration == 0:
            data1_row.append("23:59:59")
            parked_percent = 100
        else:
            parked_duration = sub_times("23:59:59", driving_duration)
            data1_row.append(parked_duration)
            parked_seconds = time.strptime(parked_duration, "%H:%M:%S")
            parked_percent = datetime.timedelta(hours=parked_seconds.tm_hour,minutes=parked_seconds.tm_min,seconds=parked_seconds.tm_sec).total_seconds() / datetime.timedelta(hours=23,minutes=59,seconds=59).total_seconds() * 100

        data1_row.append(round(parked_percent, 1))
        data1_row.append(driving_duration)
        data1_row.append(round(driving_percent, 1))
        data1_row.append(idling_duration)
        data1_row.append(round(idling_percent, 1))
        data1_row.append("23:59:59")
        
        data1.append(data1_row)

        start_date += delta

    for row_index in range(len(data1)):
        
        row_num = str(row_index + 10)
        row = data1[row_index]

        ws['A' + row_num].value = row[0]
        ws['B' + row_num].value = str(row[1])
        ws['C' + row_num].value = row[2]
        ws['D' + row_num].value = str(row[3])
        ws['E' + row_num].value = row[4]
        ws['F' + row_num].value = str(row[5])
        ws['G' + row_num].value = row[6]
        ws['H' + row_num].value = str(row[7])
        ws['I' + row_num].value = row[8]

    # create a bar chart
    category_names = ['Driving', 'Idling']

    labels = list(results.keys())
    data = np.array(list(results.values()))
    data_cum = data.cumsum(axis=1)
    category_colors = plt.get_cmap('RdYlGn')(
        np.linspace(0.15, 0.85, data.shape[1]))

    fig, ax = plt.subplots(figsize=(7.2, 5))
    ax.invert_yaxis()
    ax.xaxis.set_visible(True)
    ax.set_xlim(0, np.sum(data, axis=1).max())

    for i, (colname, color) in enumerate(zip(category_names, category_colors)):
        widths = data[:, i]
        starts = data_cum[:, i] - widths
        ax.barh(labels, widths, left=starts, height=0.5,
                label=colname, color=color)
        xcenters = starts + widths / 2

        r, g, b, _ = color
        text_color = 'white' if r * g * b < 0.5 else 'darkgrey'
        for y, (x, c) in enumerate(zip(xcenters, widths)):
            ax.text(x, y, str(int(c)), ha='center', va='center',
                    color=text_color)
    ax.legend(ncol=len(category_names), bbox_to_anchor=(0, 1),
              loc='lower left', fontsize='small')

    fig.savefig('dates_vs_duration.png')

    # add the chart to the Word document
    img = openpyxl.drawing.image.Image('dates_vs_duration.png')
    ws.add_image(img, 'A' + str(len(data1) + 11))
    # save the Word document
    wb.save(filename=save_file2)

    return save_file2

def export_data3(tracks):
    """Save and Export one file in excel worksheet and """

    wb = openpyxl.load_workbook(filename=save_file3)
    ws = wb['Sheet1']

    ws['A1'].value = 'Geofence visits'

    for row in ws.iter_rows(min_row=2, max_col=6):
        for cell in row:
            cell.value = None

    ws['A2'].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + ' - intervals:17'

    ws['A4'].value = 'Geofence'
    ws['B4'].value = 'Entrance_Time'
    ws['C4'].value = 'Entrance_Place'
    ws['D4'].value = 'Exit_Time'
    ws['E4'].value = 'Exit_Place'
    ws['F4'].value = 'Duration'

    for row_index in range(len(tracks)):
        row_num = str(row_index + 5)
        row = tracks[row_index]

        ws['A' + row_num].value = row['track_label']
        ws['B' + row_num].value = row['time_start']
        ws['C' + row_num].value = row['from_address']
        ws['D' + row_num].value = row['time_end']
        ws['E' + row_num].value = row['to_address']
        ws['F' + row_num].value = row['duration']

    wb.save(filename=save_file3)

    return save_file3

def export_data4(tracks):

    """Save and Export one file in excel worksheet and """

    wb = openpyxl.load_workbook(filename=save_file4)
    ws = wb['Sheet1']

    for row in ws.iter_rows(min_row=1, max_col=10):
        for cell in row:
            cell.value = None

    ws['A1'].value = 'KAMOTO COPPER COMPANY OFFLINE DEVICES'
    ws['C1'].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws['A3'].value = 'Clients: KAMOTO COPPER COMPANY'
    ws['A4'].value = 'Max Hours: 96'
    ws['A5'].value = 'Schedule: VEHICLES OFFLINE REPORT'
    ws['A6'].value = 'KAMOTO COPPER COMPANY'

    ws['A8'].value = 'Device'
    ws['B8'].value = 'Asset'
    ws['C8'].value = 'Last Date'
    ws['D8'].value = 'Last Update'
    ws['E8'].value = 'Location'
    ws['F8'].value = 'Remarque'

    row_num = 0

    for row_index in range(len(tracks)):
        row = tracks[row_index]

        last_update = datetime.datetime.strptime(row['time_end'], "%Y-%m-%d %H:%M:%S").date()
        today = datetime.datetime.today().date()

        days_ago = (today - last_update).days

        if days_ago > 0:
            time_ago = str(days_ago) + ' days ago'
            if days_ago > 6:
                week_ago = int((today - last_update).days / 7)
                time_ago = str(week_ago) + ' weeks ago'        

            row_num = row_num + 1

            ws['A' + str(row_num)].value = row['track_id']
            ws['B' + str(row_num)].value = row['track_label']
            ws['C' + str(row_num)].value = row['time_end']
            ws['D' + str(row_num)].value = time_ago
            ws['E' + str(row_num)].value = row['to_address']
            ws['F' + str(row_num)].value = row['duration']

    # create a table from the data
    max_row=8+len(tracks)

    table = Table(displayName=datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"), ref='A8:F'+str(max_row))

    # add a style to the table
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)

    # add sort buttons to the table
    style.showColumnStripes = False
    style.showFirstColumn = True
    style.showLastColumn = True

    table.tableStyleInfo = style

    # add the table to the worksheet
    ws.add_table(table)

    wb.save(filename=save_file4)

    return save_file4

def sum_times(time1, time2):
    # Convert time strings to seconds
    time1_secs = sum(int(x) * 60 ** i for i, x in enumerate(reversed(time1.split(":"))))
    time2_secs = sum(int(x) * 60 ** i for i, x in enumerate(reversed(time2.split(":"))))
    
    # Calculate total seconds
    total_secs = time1_secs + time2_secs
    
    # Convert total seconds back to "hh:mm:ss" format
    hours, remainder = divmod(total_secs, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def sub_times(time1, time2):
    # Convert time strings to seconds
    time1_secs = sum(int(x) * 60 ** i for i, x in enumerate(reversed(time1.split(":"))))
    time2_secs = sum(int(x) * 60 ** i for i, x in enumerate(reversed(time2.split(":"))))
    
    # Calculate total seconds
    sub_secs = time1_secs - time2_secs
    
    # Convert total seconds back to "hh:mm:ss" format
    hours, remainder = divmod(sub_secs, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"