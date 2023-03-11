import os
import openpyxl

APP_ROOT = os.path.dirname(os.path.abspath(__file__))   # refers to application_top
APP_SAVE = os.path.join(APP_ROOT, '../download/data.xlsx')

save_file = APP_SAVE

def export_data(tracks):
    """Save and Export one file in excel worksheet and """

    wb = openpyxl.load_workbook(filename=save_file)
    ws = wb['Sheet1']

    for row in ws.iter_rows(min_row=2, max_col=10):
        for cell in row:
            cell.value = None

    for row_index in range(len(tracks)):
        row_num = str(row_index + 2)
        row = tracks[row_index]

        ws['A' + row_num].value = row['record_id']
        ws['B' + row_num].value = row['track_id']
        ws['C' + row_num].value = row['track_label']
        ws['D' + row_num].value = row['time_start']
        ws['E' + row_num].value = row['time_end']
        ws['F' + row_num].value = row['from_address']
        ws['G' + row_num].value = row['to_address']
        ws['H' + row_num].value = row['distance']
        ws['I' + row_num].value = row['duration']
        ws['J' + row_num].value = row['max_speed']
        ws['K' + row_num].value = row['Driver']
        ws['L' + row_num].value = row['TT_number_tags']
        ws['M' + row_num].value = row['Registration_plate']
        ws['N' + row_num].value = row['Product']
        ws['O' + row_num].value = row['Parked']
        ws['P' + row_num].value = row['Idle_time']
        ws['Q' + row_num].value = row['status']
        ws['R' + row_num].value = row['zone']

    wb.save(filename=save_file)

    return save_file